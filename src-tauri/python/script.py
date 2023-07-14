import openpyxl
import json
import sys


wb = openpyxl.load_workbook(sys.argv[1])
sheetnames = wb.sheetnames
ws = wb[sheetnames[0]]
# if End col is specified / not None stop looping after this col other wise read the whole thing
end_col = None

row_data = {'headers': ['ID'], 'data': {}}

# Find the column index for the header you want
header_name = 'Note'
header_col_index = None
for col in ws.iter_cols():
    row_data['headers'].append(col[0].value)
    if col[0].value == header_name:
        header_col_index = col[0].column
        break


for row in ws.iter_rows(min_row=2, max_col=header_col_index):
    cell_data = {}
    for cell in row:
       
        cell_dict = {'value': cell.value, 'url': ""}
        if cell.hyperlink is not None:
            cell_dict['url'] = cell.hyperlink.target
        cell_data[row_data['headers'][cell.column]] = cell_dict
        cell_data['ID'] = cell.row-1 #same as key because when object will be transformed in array, and key will be lost if not here
    row_data['data'][cell.row-1] = cell_data

# with open('cell_values.json', 'w') as outfile:
json.dump(row_data, sys.stdout)


    
